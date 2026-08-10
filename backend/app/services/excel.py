from __future__ import annotations

import hashlib
import os
import posixpath
import re
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from html import escape
from pathlib import Path
from typing import Any, BinaryIO, Dict, Iterable, List, Mapping, Optional, Sequence
from uuid import uuid4
from xml.etree import ElementTree
from zipfile import BadZipFile, ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

REQUIRED_HEADERS = (
    "상품코드",
    "상품명",
    "규격",
    "현재고",
    "매입단가",
    "공급사코드",
    "공급사",
)
TARGET_SHEET_NAME = "Sheet"
SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOCUMENT_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
WORKSHEET_REL_TYPE = f"{DOCUMENT_REL_NS}/worksheet"
WORKSHEET_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
)


class ExcelValidationError(ValueError):
    pass


@dataclass(frozen=True)
class StagedExcel:
    path: Path
    storage_name: str
    original_name: str
    sha256: str


@dataclass(frozen=True)
class ProductRecord:
    product_code: str
    product_name: Optional[str]
    specification: Optional[str]
    current_stock: Any
    purchase_price: Any
    supplier_code: Optional[str]
    supplier: Optional[str]
    excel_row: int


@dataclass(frozen=True)
class InventoryCellUpdate:
    excel_row: int
    expected_product_code: str
    current_stock: Optional[Any] = None
    purchase_price: Optional[Any] = None


@dataclass(frozen=True)
class RegisteredProductRow:
    excel_row: int
    product_code: str
    product_name: str
    specification: Optional[str]
    current_stock: int
    purchase_price: Optional[int]
    supplier_code: Optional[str]
    supplier: Optional[str]


@dataclass(frozen=True)
class HistorySheetAppend:
    sheet_name: str
    headers: Sequence[str]
    rows: Sequence[Sequence[Any]]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file_handle:
        for chunk in iter(lambda: file_handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def stage_excel_upload(
    source: BinaryIO, original_filename: str, uploads_dir: Path
) -> StagedExcel:
    original_name = original_filename.replace("\\", "/").rsplit("/", 1)[-1]
    if not original_name or Path(original_name).suffix.lower() != ".xlsx":
        raise ExcelValidationError(".xlsx 형식의 상품리스트만 업로드할 수 있습니다.")

    uploads_dir.mkdir(parents=True, exist_ok=True)
    storage_name = f"{uuid4()}.xlsx"
    staged_path = uploads_dir / f".{storage_name}"
    digest = hashlib.sha256()

    try:
        source.seek(0)
        with staged_path.open("xb") as destination:
            while chunk := source.read(1024 * 1024):
                digest.update(chunk)
                destination.write(chunk)
    except Exception:
        staged_path.unlink(missing_ok=True)
        raise

    if staged_path.stat().st_size == 0:
        staged_path.unlink(missing_ok=True)
        raise ExcelValidationError("빈 Excel 파일은 업로드할 수 없습니다.")

    return StagedExcel(
        path=staged_path,
        storage_name=storage_name,
        original_name=original_name,
        sha256=digest.hexdigest(),
    )


def validate_product_workbook(path: Path) -> List[ProductRecord]:
    try:
        with path.open("rb") as file_handle:
            workbook = load_workbook(file_handle, read_only=False, data_only=False)
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError, EOFError) as exc:
        raise ExcelValidationError(
            "손상되었거나 암호화된 Excel 파일은 사용할 수 없습니다."
        ) from exc

    try:
        if TARGET_SHEET_NAME not in workbook.sheetnames:
            raise ExcelValidationError("'Sheet' 시트를 찾을 수 없습니다.")

        worksheet = workbook[TARGET_SHEET_NAME]
        header_positions = _required_header_positions(worksheet)
        records: List[ProductRecord] = []
        seen_codes: Dict[str, int] = {}

        for row_number in range(2, worksheet.max_row + 1):
            values = {
                header: worksheet.cell(row=row_number, column=column).value
                for header, column in header_positions.items()
            }
            if _row_is_empty(values.values()):
                continue

            product_code = _normalize_product_code(values["상품코드"])
            if not product_code:
                raise ExcelValidationError(
                    f"{row_number}행의 상품코드가 비어 있습니다."
                )
            if product_code in seen_codes:
                first_row = seen_codes[product_code]
                raise ExcelValidationError(
                    f"상품코드 '{product_code}'가 {first_row}행과 "
                    f"{row_number}행에 중복되어 있습니다."
                )
            seen_codes[product_code] = row_number

            purchase_price = values["매입단가"]
            if _is_negative_number(purchase_price):
                raise ExcelValidationError(
                    f"{row_number}행의 매입단가는 0 이상이어야 합니다."
                )

            records.append(
                ProductRecord(
                    product_code=product_code,
                    product_name=_optional_text(values["상품명"]),
                    specification=_optional_text(values["규격"]),
                    current_stock=_json_cell_value(values["현재고"]),
                    purchase_price=_json_cell_value(purchase_price),
                    supplier_code=_optional_text(values["공급사코드"]),
                    supplier=_optional_text(values["공급사"]),
                    excel_row=row_number,
                )
            )

        return records
    finally:
        workbook.close()


def product_sheet_max_row(path: Path) -> int:
    try:
        workbook = load_workbook(path, read_only=False, data_only=False)
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError, EOFError) as exc:
        raise ExcelValidationError(
            "손상되었거나 암호화된 Excel 파일은 사용할 수 없습니다."
        ) from exc
    try:
        if TARGET_SHEET_NAME not in workbook.sheetnames:
            raise ExcelValidationError("'Sheet' 시트를 찾을 수 없습니다.")
        return max(1, workbook[TARGET_SHEET_NAME].max_row)
    finally:
        workbook.close()


def create_inventory_copy(
    source_path: Path,
    destination_path: Path,
    updates: Sequence[InventoryCellUpdate],
    history_sheets: Sequence[HistorySheetAppend] = (),
    registered_products: Sequence[RegisteredProductRow] = (),
) -> None:
    if source_path.resolve() == destination_path.resolve():
        raise ValueError("원본 Excel과 출력 Excel 경로는 달라야 합니다.")

    original_hash = sha256_file(source_path)
    try:
        workbook = load_workbook(source_path, read_only=False, data_only=False)
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError, EOFError) as exc:
        raise ExcelValidationError(
            "손상되었거나 암호화된 Excel 파일은 사용할 수 없습니다."
        ) from exc

    temporary_path = destination_path.with_name(
        f".{destination_path.name}.{uuid4().hex}.tmp.xlsx"
    )
    try:
        if TARGET_SHEET_NAME not in workbook.sheetnames:
            raise ExcelValidationError("'Sheet' 시트를 찾을 수 없습니다.")
        worksheet = workbook[TARGET_SHEET_NAME]
        header_positions = _required_header_positions(worksheet)
        sheet_paths = _workbook_sheet_paths(source_path)
        replacements: Dict[str, bytes] = {}
        new_entries: Dict[str, bytes] = {}

        with ZipFile(source_path, "r") as source_archive:
            target_sheet_path = sheet_paths[TARGET_SHEET_NAME]
            target_sheet_xml = source_archive.read(target_sheet_path)

            if registered_products:
                target_sheet_xml = _append_registered_products(
                    target_sheet_xml,
                    registered_products,
                    header_positions,
                    worksheet,
                )

            for update in updates:
                if update.excel_row < 2 or update.excel_row > worksheet.max_row:
                    raise ExcelValidationError(
                        f"{update.excel_row}행은 수정할 수 있는 상품 행이 아닙니다."
                    )
                actual_product_code = _normalize_product_code(
                    worksheet.cell(
                        row=update.excel_row,
                        column=header_positions["상품코드"],
                    ).value
                )
                if actual_product_code != _normalize_product_code(
                    update.expected_product_code
                ):
                    raise ExcelValidationError(
                        f"{update.excel_row}행의 상품코드가 예상값과 다릅니다."
                    )
                if update.current_stock is not None:
                    value = _integer_value(update.current_stock, "현재고", allow_negative=True)
                    coordinate = (
                        f"{get_column_letter(header_positions['현재고'])}{update.excel_row}"
                    )
                    target_sheet_xml = _set_numeric_cell(
                        target_sheet_xml, coordinate, value
                    )
                if update.purchase_price is not None:
                    value = _integer_value(
                        update.purchase_price, "매입단가", allow_negative=False
                    )
                    coordinate = (
                        f"{get_column_letter(header_positions['매입단가'])}{update.excel_row}"
                    )
                    target_sheet_xml = _set_numeric_cell(
                        target_sheet_xml, coordinate, value
                    )

            replacements[target_sheet_path] = target_sheet_xml

            seen_history_names: set[str] = set()
            for history in history_sheets:
                _validate_history_append(history, seen_history_names)
                if history.sheet_name in workbook.sheetnames:
                    history_sheet = workbook[history.sheet_name]
                    existing_header_values = [
                        history_sheet.cell(row=1, column=column).value
                        for column in range(
                            1,
                            max(history_sheet.max_column, len(history.headers)) + 1,
                        )
                    ]
                    while existing_header_values and (
                        existing_header_values[-1] is None
                        or (
                            isinstance(existing_header_values[-1], str)
                            and not existing_header_values[-1].strip()
                        )
                    ):
                        existing_header_values.pop()
                    existing_headers = tuple(existing_header_values)
                    if existing_headers != tuple(history.headers):
                        raise ExcelValidationError(
                            f"'{history.sheet_name}' 시트의 기존 헤더가 호환되지 않습니다."
                        )
                    sheet_path = sheet_paths[history.sheet_name]
                    sheet_xml = replacements.get(
                        sheet_path, source_archive.read(sheet_path)
                    )
                    replacements[sheet_path] = _append_history_rows(
                        sheet_xml,
                        history_sheet.max_row + 1,
                        history.rows,
                        max(history_sheet.max_column, len(history.headers)),
                    )
                else:
                    manifest_replacements, sheet_path = _add_sheet_to_manifest(
                        source_archive,
                        replacements,
                        history.sheet_name,
                    )
                    replacements.update(manifest_replacements)
                    new_entries[sheet_path] = _new_history_sheet_xml(
                        history.headers, history.rows
                    )

            destination_path.parent.mkdir(parents=True, exist_ok=True)
            with ZipFile(temporary_path, "x", compression=ZIP_DEFLATED) as output_archive:
                for entry in source_archive.infolist():
                    output_archive.writestr(
                        entry,
                        replacements.get(entry.filename, source_archive.read(entry.filename)),
                    )
                for filename, content in new_entries.items():
                    output_archive.writestr(filename, content)

        workbook.close()
        workbook = None

        _verify_inventory_updates(
            temporary_path, updates, history_sheets, registered_products
        )
        if sha256_file(source_path) != original_hash:
            raise RuntimeError("원본 Excel 파일이 변경되어 출력을 중단했습니다.")
        os.replace(temporary_path, destination_path)
    finally:
        if workbook is not None:
            workbook.close()
        temporary_path.unlink(missing_ok=True)


def _required_header_positions(worksheet: Any) -> Dict[str, int]:
    positions: Dict[str, List[int]] = {header: [] for header in REQUIRED_HEADERS}
    for column, cell in enumerate(worksheet[1], start=1):
        header = cell.value.strip() if isinstance(cell.value, str) else cell.value
        if header in positions:
            positions[header].append(column)

    missing = [header for header, columns in positions.items() if not columns]
    if missing:
        raise ExcelValidationError(
            "필수 헤더가 누락되었습니다: " + ", ".join(missing)
        )

    duplicated = [header for header, columns in positions.items() if len(columns) > 1]
    if duplicated:
        raise ExcelValidationError(
            "필수 헤더가 중복되었습니다: " + ", ".join(duplicated)
        )

    return {header: columns[0] for header, columns in positions.items()}


def _verify_inventory_updates(
    workbook_path: Path,
    updates: Sequence[InventoryCellUpdate],
    history_sheets: Sequence[HistorySheetAppend],
    registered_products: Sequence[RegisteredProductRow],
) -> None:
    try:
        workbook = load_workbook(workbook_path, read_only=False, data_only=False)
        worksheet = workbook[TARGET_SHEET_NAME]
        header_positions = _required_header_positions(worksheet)
        for update in updates:
            actual_product_code = _normalize_product_code(
                worksheet.cell(
                    row=update.excel_row,
                    column=header_positions["상품코드"],
                ).value
            )
            if actual_product_code != _normalize_product_code(
                update.expected_product_code
            ):
                raise RuntimeError("상품코드 출력값을 검증할 수 없습니다.")
            if update.current_stock is not None:
                actual_stock = worksheet.cell(
                    row=update.excel_row, column=header_positions["현재고"]
                ).value
                if actual_stock != update.current_stock:
                    raise RuntimeError("현재고 출력값을 검증할 수 없습니다.")
            if update.purchase_price is not None:
                actual_price = worksheet.cell(
                    row=update.excel_row, column=header_positions["매입단가"]
                ).value
                if actual_price != update.purchase_price:
                    raise RuntimeError("매입단가 출력값을 검증할 수 없습니다.")
        for product in registered_products:
            expected = (
                product.product_code,
                product.product_name,
                product.specification,
                product.current_stock,
                product.purchase_price,
                product.supplier_code,
                product.supplier,
            )
            actual = tuple(
                worksheet.cell(
                    row=product.excel_row,
                    column=header_positions[header],
                ).value
                for header in REQUIRED_HEADERS
            )
            if actual != expected:
                raise RuntimeError("신규 상품 출력값을 검증할 수 없습니다.")
        for history in history_sheets:
            history_sheet = workbook[history.sheet_name]
            if history.rows:
                first_row = history_sheet.max_row - len(history.rows) + 1
                actual_rows = tuple(
                    tuple(
                        history_sheet.cell(row=row, column=column).value
                        for column in range(1, len(history.headers) + 1)
                    )
                    for row in range(first_row, history_sheet.max_row + 1)
                )
                if actual_rows != tuple(tuple(row) for row in history.rows):
                    raise RuntimeError("이력 시트 출력값을 검증할 수 없습니다.")
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError, EOFError) as exc:
        raise RuntimeError("생성한 Excel 파일을 다시 열어 검증할 수 없습니다.") from exc
    finally:
        if "workbook" in locals():
            workbook.close()


def _workbook_sheet_paths(workbook_path: Path) -> Dict[str, str]:
    try:
        with ZipFile(workbook_path, "r") as archive:
            workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
            relationships_root = ElementTree.fromstring(
                archive.read("xl/_rels/workbook.xml.rels")
            )
    except (BadZipFile, KeyError, ElementTree.ParseError) as exc:
        raise ExcelValidationError("Excel 내부 시트 구조를 확인할 수 없습니다.") from exc

    relationship_targets = {
        relationship.attrib["Id"]: relationship.attrib["Target"]
        for relationship in relationships_root.findall(f"{{{PACKAGE_REL_NS}}}Relationship")
        if relationship.attrib.get("Type") == WORKSHEET_REL_TYPE
    }
    paths: Dict[str, str] = {}
    for sheet in workbook_root.findall(f".//{{{SPREADSHEET_NS}}}sheet"):
        relationship_id = sheet.attrib.get(f"{{{DOCUMENT_REL_NS}}}id")
        target = relationship_targets.get(relationship_id or "")
        if not target:
            continue
        paths[sheet.attrib["name"]] = (
            target.lstrip("/")
            if target.startswith("/")
            else posixpath.normpath(posixpath.join("xl", target))
        )
    return paths


def _integer_value(value: Any, label: str, allow_negative: bool) -> int:
    if isinstance(value, bool) or not isinstance(value, int):
        raise ExcelValidationError(f"{label}는 정수여야 합니다.")
    if not allow_negative and value < 0:
        raise ExcelValidationError(f"{label}는 0 이상이어야 합니다.")
    return value


def _append_registered_products(
    sheet_xml: bytes,
    products: Sequence[RegisteredProductRow],
    header_positions: Mapping[str, int],
    worksheet: Any,
) -> bytes:
    seen_rows: set[int] = set()
    seen_codes: set[str] = set()
    source_max_row = worksheet.max_row
    style_row = source_max_row if source_max_row >= 2 else None
    styles = {
        header: (
            worksheet.cell(row=style_row, column=column).style_id
            if style_row is not None
            else 0
        )
        for header, column in header_positions.items()
    }
    rows: list[bytes] = []
    for product in sorted(products, key=lambda row: row.excel_row):
        if product.excel_row <= source_max_row or product.excel_row in seen_rows:
            raise ExcelValidationError("신규 상품 Excel 행이 올바르지 않습니다.")
        normalized_code = _normalize_product_code(product.product_code)
        if not normalized_code or normalized_code in seen_codes:
            raise ExcelValidationError("신규 상품코드가 비어 있거나 중복되었습니다.")
        seen_rows.add(product.excel_row)
        seen_codes.add(normalized_code)
        values = {
            "상품코드": product.product_code,
            "상품명": product.product_name,
            "규격": product.specification,
            "현재고": _integer_value(product.current_stock, "현재고", True),
            "매입단가": (
                None
                if product.purchase_price is None
                else _integer_value(product.purchase_price, "매입단가", False)
            ),
            "공급사코드": product.supplier_code,
            "공급사": product.supplier,
        }
        cells = b"".join(
            _styled_cell_xml(
                get_column_letter(header_positions[header]),
                product.excel_row,
                values[header],
                styles[header],
            )
            for header in REQUIRED_HEADERS
        )
        rows.append(
            b'<row r="'
            + str(product.excel_row).encode("ascii")
            + b'">'
            + cells
            + b"</row>"
        )
    row_xml = b"".join(rows)
    if b"</sheetData>" in sheet_xml:
        sheet_xml = sheet_xml.replace(b"</sheetData>", row_xml + b"</sheetData>", 1)
    elif b"<sheetData/>" in sheet_xml:
        sheet_xml = sheet_xml.replace(
            b"<sheetData/>", b"<sheetData>" + row_xml + b"</sheetData>", 1
        )
    else:
        raise ExcelValidationError("상품 시트의 행 구조를 확인할 수 없습니다.")
    return _update_dimension(
        sheet_xml,
        max(worksheet.max_column, *header_positions.values()),
        max(seen_rows),
    )


def _styled_cell_xml(
    column: str, row_number: int, value: Any, style_id: int
) -> bytes:
    cell = _history_cell_xml(column, row_number, value)
    if style_id:
        return cell.replace(
            b"<c ", f'<c s="{style_id}" '.encode("ascii"), 1
        )
    return cell


def _set_numeric_cell(sheet_xml: bytes, coordinate: str, value: int) -> bytes:
    coordinate_bytes = coordinate.encode("ascii")
    cell_pattern = re.compile(
        rb'<c\b(?=[^>]*\br=(["\'])'
        + re.escape(coordinate_bytes)
        + rb'\1)[^>]*(?:/>|>.*?</c>)',
        re.DOTALL,
    )
    match = cell_pattern.search(sheet_xml)
    numeric_xml = str(value).encode("ascii")
    if match:
        cell_xml = match.group(0)
        opening_end = cell_xml.find(b">")
        opening = cell_xml[: opening_end + 1]
        if opening.endswith(b"/>"):
            opening = opening[:-2] + b">"
        opening = re.sub(rb'\s+t=(["\']).*?\1', b"", opening)
        replacement = opening + b"<v>" + numeric_xml + b"</v></c>"
        return sheet_xml[: match.start()] + replacement + sheet_xml[match.end() :]

    row_number = re.search(rb"\d+$", coordinate_bytes)
    if row_number is None:
        raise ExcelValidationError("수정할 셀 주소가 올바르지 않습니다.")
    row_pattern = re.compile(
        rb'<row\b(?=[^>]*\br=(["\'])'
        + row_number.group(0)
        + rb'\1)[^>]*>.*?</row>',
        re.DOTALL,
    )
    row_match = row_pattern.search(sheet_xml)
    if row_match is None:
        raise ExcelValidationError(f"{coordinate} 셀의 행을 찾을 수 없습니다.")
    cell = b'<c r="' + coordinate_bytes + b'"><v>' + numeric_xml + b"</v></c>"
    row_xml = row_match.group(0)
    target_column = column_index_from_string(
        re.match(rb"[A-Z]+", coordinate_bytes).group(0).decode("ascii")  # type: ignore[union-attr]
    )
    insertion = row_xml.rfind(b"</row>")
    existing_cell_pattern = re.compile(
        rb'<c\b[^>]*\br=["\']([A-Z]+)\d+["\'][^>]*(?:/>|>.*?</c>)',
        re.DOTALL,
    )
    for existing_cell in existing_cell_pattern.finditer(row_xml):
        existing_column = column_index_from_string(
            existing_cell.group(1).decode("ascii")
        )
        if existing_column > target_column:
            insertion = existing_cell.start()
            break
    patched_row = row_xml[:insertion] + cell + row_xml[insertion:]
    return sheet_xml[: row_match.start()] + patched_row + sheet_xml[row_match.end() :]


def _validate_history_append(
    history: HistorySheetAppend, seen_history_names: set[str]
) -> None:
    if (
        not history.sheet_name
        or len(history.sheet_name) > 31
        or any(character in history.sheet_name for character in "[]:*?/\\")
    ):
        raise ExcelValidationError("이력 시트 이름이 올바르지 않습니다.")
    if history.sheet_name in seen_history_names:
        raise ExcelValidationError("같은 이력 시트를 한 번만 지정해 주세요.")
    seen_history_names.add(history.sheet_name)
    if not history.headers or any(not header for header in history.headers):
        raise ExcelValidationError("이력 시트 헤더가 올바르지 않습니다.")
    for row in history.rows:
        if len(row) != len(history.headers):
            raise ExcelValidationError(
                f"'{history.sheet_name}' 이력 행의 열 수가 헤더와 다릅니다."
            )


def _append_history_rows(
    sheet_xml: bytes,
    start_row: int,
    rows: Sequence[Sequence[Any]],
    dimension_columns: int,
) -> bytes:
    if not rows:
        return sheet_xml
    row_xml = b"".join(
        _history_row_xml(start_row + offset, row)
        for offset, row in enumerate(rows)
    )
    if b"</sheetData>" in sheet_xml:
        sheet_xml = sheet_xml.replace(b"</sheetData>", row_xml + b"</sheetData>", 1)
    elif b"<sheetData/>" in sheet_xml:
        sheet_xml = sheet_xml.replace(
            b"<sheetData/>", b"<sheetData>" + row_xml + b"</sheetData>", 1
        )
    else:
        raise ExcelValidationError("이력 시트의 행 구조를 확인할 수 없습니다.")
    return _update_dimension(
        sheet_xml, dimension_columns, start_row + len(rows) - 1
    )


def _history_row_xml(row_number: int, values: Sequence[Any]) -> bytes:
    cells = b"".join(
        _history_cell_xml(get_column_letter(column), row_number, value)
        for column, value in enumerate(values, start=1)
    )
    return b'<row r="' + str(row_number).encode("ascii") + b'">' + cells + b"</row>"


def _history_cell_xml(column: str, row_number: int, value: Any) -> bytes:
    coordinate = f"{column}{row_number}"
    if value is None:
        return f'<c r="{coordinate}"/>'.encode("utf-8")
    if isinstance(value, bool):
        return f'<c r="{coordinate}" t="b"><v>{int(value)}</v></c>'.encode("utf-8")
    if isinstance(value, (int, float, Decimal)):
        return f'<c r="{coordinate}"><v>{value}</v></c>'.encode("utf-8")
    if isinstance(value, (date, datetime, time)):
        value = value.isoformat()
    text_value = str(value)
    preserve_space = ' xml:space="preserve"' if text_value != text_value.strip() else ""
    return (
        f'<c r="{coordinate}" t="inlineStr"><is><t{preserve_space}>'
        f"{escape(text_value)}</t></is></c>"
    ).encode("utf-8")


def _new_history_sheet_xml(
    headers: Sequence[str], rows: Sequence[Sequence[Any]]
) -> bytes:
    all_rows: Sequence[Sequence[Any]] = (tuple(headers), *tuple(tuple(row) for row in rows))
    sheet_rows = b"".join(
        _history_row_xml(row_number, row)
        for row_number, row in enumerate(all_rows, start=1)
    )
    dimension = f"A1:{get_column_letter(len(headers))}{len(all_rows)}"
    return (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        + f'<worksheet xmlns="{SPREADSHEET_NS}"><dimension ref="{dimension}"/>'
        .encode("utf-8")
        + b"<sheetData>"
        + sheet_rows
        + b"</sheetData></worksheet>"
    )


def _update_dimension(sheet_xml: bytes, columns: int, last_row: int) -> bytes:
    reference = f'A1:{get_column_letter(columns)}{last_row}'.encode("ascii")
    dimension_pattern = re.compile(rb'(<dimension\b[^>]*\bref=")[^"]*(")')
    if dimension_pattern.search(sheet_xml):
        return dimension_pattern.sub(rb"\g<1>" + reference + rb"\g<2>", sheet_xml, count=1)
    return sheet_xml


def _add_sheet_to_manifest(
    source_archive: ZipFile,
    current_replacements: Mapping[str, bytes],
    sheet_name: str,
) -> tuple[Dict[str, bytes], str]:
    workbook_path = "xl/workbook.xml"
    relationships_path = "xl/_rels/workbook.xml.rels"
    content_types_path = "[Content_Types].xml"
    workbook_xml = current_replacements.get(
        workbook_path, source_archive.read(workbook_path)
    )
    relationships_xml = current_replacements.get(
        relationships_path, source_archive.read(relationships_path)
    )
    content_types_xml = current_replacements.get(
        content_types_path, source_archive.read(content_types_path)
    )

    workbook_root = ElementTree.fromstring(workbook_xml)
    relationship_root = ElementTree.fromstring(relationships_xml)
    sheet_ids = [
        int(sheet.attrib["sheetId"])
        for sheet in workbook_root.findall(f".//{{{SPREADSHEET_NS}}}sheet")
    ]
    relationship_ids = {
        relationship.attrib["Id"]
        for relationship in relationship_root.findall(f"{{{PACKAGE_REL_NS}}}Relationship")
    }
    relationship_number = 1
    while f"rId{relationship_number}" in relationship_ids:
        relationship_number += 1
    relationship_id = f"rId{relationship_number}"

    used_sheet_parts = set(source_archive.namelist())
    sheet_number = 1
    while (
        f"xl/worksheets/sheet{sheet_number}.xml" in used_sheet_parts
        or f'/xl/worksheets/sheet{sheet_number}.xml'.encode("ascii")
        in content_types_xml
    ):
        sheet_number += 1
    sheet_path = f"xl/worksheets/sheet{sheet_number}.xml"

    sheet_fragment = (
        f'<sheet name="{escape(sheet_name, quote=True)}" '
        f'sheetId="{max(sheet_ids, default=0) + 1}" r:id="{relationship_id}"/>'
    ).encode("utf-8")
    relationship_fragment = (
        f'<Relationship Id="{relationship_id}" Type="{WORKSHEET_REL_TYPE}" '
        f'Target="worksheets/sheet{sheet_number}.xml"/>'
    ).encode("utf-8")
    override_fragment = (
        f'<Override PartName="/xl/worksheets/sheet{sheet_number}.xml" '
        f'ContentType="{WORKSHEET_CONTENT_TYPE}"/>'
    ).encode("utf-8")

    return (
        {
            workbook_path: _insert_before_closing(
                workbook_xml, b"</sheets>", sheet_fragment
            ),
            relationships_path: _insert_before_closing(
                relationships_xml, b"</Relationships>", relationship_fragment
            ),
            content_types_path: _insert_before_closing(
                content_types_xml, b"</Types>", override_fragment
            ),
        },
        sheet_path,
    )


def _insert_before_closing(xml: bytes, closing: bytes, fragment: bytes) -> bytes:
    position = xml.rfind(closing)
    if position < 0:
        raise ExcelValidationError("Excel 내부 XML 구조를 갱신할 수 없습니다.")
    return xml[:position] + fragment + xml[position:]


def _row_is_empty(values: Iterable[Any]) -> bool:
    return all(value is None or (isinstance(value, str) and not value.strip()) for value in values)


def _normalize_product_code(value: Any) -> str:
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _optional_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def _json_cell_value(value: Any) -> Any:
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _is_negative_number(value: Any) -> bool:
    if value is None or isinstance(value, bool):
        return False
    if isinstance(value, (int, float, Decimal)):
        return value < 0
    if isinstance(value, str):
        normalized = value.replace(",", "").strip()
        try:
            return bool(normalized) and Decimal(normalized) < 0
        except ArithmeticError:
            return False
    return False
