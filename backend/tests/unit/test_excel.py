from pathlib import Path
import re
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.services.excel import (
    ExcelValidationError,
    HistorySheetAppend,
    InventoryCellUpdate,
    create_inventory_copy,
    sha256_file,
    validate_product_workbook,
)

HEADERS = [
    "상품코드",
    "상품명",
    "규격",
    "현재고",
    "매입단가",
    "공급사코드",
    "공급사",
]


def save_workbook(path: Path, headers=None, rows=None) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet"
    worksheet.append(headers or HEADERS)
    for row in rows or []:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def test_optional_duplicate_header_is_allowed(tmp_path: Path) -> None:
    path = tmp_path / "products.xlsx"
    headers = HEADERS + ["상품유형", "상품유형"]
    save_workbook(path, headers, [["0001", "상품", "1정", 3, 100, "S1", "공급사", "A", "B"]])

    records = validate_product_workbook(path)

    assert len(records) == 1
    assert records[0].product_code == "0001"
    assert records[0].excel_row == 2


@pytest.mark.parametrize(
    ("headers", "rows", "message"),
    [
        (HEADERS[:-1], [["1", "상품", "1정", 0, 10, "S1"]], "필수 헤더가 누락"),
        (HEADERS + ["상품코드"], [], "필수 헤더가 중복"),
        (HEADERS, [[" ", "상품", "1정", 0, 10, "S1", "공급사"]], "상품코드가 비어"),
        (
            HEADERS,
            [
                ["DUP", "상품1", "1정", 0, 10, "S1", "공급사"],
                [" DUP ", "상품2", "1정", 0, 10, "S1", "공급사"],
            ],
            "중복",
        ),
        (
            HEADERS,
            [["0001", "상품", "1정", 0, -1, "S1", "공급사"]],
            "매입단가는 0 이상",
        ),
    ],
)
def test_invalid_workbook_is_rejected(
    tmp_path: Path, headers, rows, message: str
) -> None:
    path = tmp_path / "invalid.xlsx"
    save_workbook(path, headers, rows)

    with pytest.raises(ExcelValidationError, match=message):
        validate_product_workbook(path)


def test_inventory_copy_changes_only_target_cells(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet"
    worksheet.append(HEADERS + ["계산식", "메모"])
    worksheet.append(["0001", "상품", "1정", 10, 100, "S1", "공급사", "=D2+1", "유지"])
    worksheet["A2"].font = Font(bold=True, color="FF0000")
    worksheet["I2"].fill = PatternFill(fill_type="solid", fgColor="00FF00")
    worksheet["I3"] = ""
    source_formula = worksheet["H2"].value
    source_a_style = worksheet["A2"].style_id
    source_i_style = worksheet["I2"].style_id
    workbook.save(source)
    workbook.close()
    source_hash = sha256_file(source)

    create_inventory_copy(
        source,
        destination,
        [
            InventoryCellUpdate(
                excel_row=2,
                expected_product_code="0001",
                current_stock=15,
                purchase_price=120,
            )
        ],
        [
            HistorySheetAppend(
                sheet_name="입고반영내역",
                headers=("작업 ID", "상품코드", "변경 후 재고"),
                rows=(("job-1", "0001", 15),),
            )
        ],
    )

    assert sha256_file(source) == source_hash
    original = load_workbook(source, data_only=False)
    result = load_workbook(destination, data_only=False)
    assert original["Sheet"]["D2"].value == 10
    assert original["Sheet"]["E2"].value == 100
    assert result["Sheet"]["D2"].value == 15
    assert result["Sheet"]["E2"].value == 120
    assert result["Sheet"]["A2"].value == "0001"
    assert result["Sheet"]["H2"].value == source_formula
    assert result["Sheet"]["I2"].value == "유지"
    assert result["Sheet"]["A2"].style_id == source_a_style
    assert result["Sheet"]["I2"].style_id == source_i_style
    assert "입고반영내역" not in original.sheetnames
    assert list(result["입고반영내역"].values) == [
        ("작업 ID", "상품코드", "변경 후 재고"),
        ("job-1", "0001", 15),
    ]
    original.close()
    result.close()

    with ZipFile(source) as original_archive, ZipFile(destination) as result_archive:
        original_sheet = original_archive.read("xl/worksheets/sheet1.xml")
        result_sheet = result_archive.read("xl/worksheets/sheet1.xml")
        cell_pattern = re.compile(
            rb'<c\b[^>]*\br="([A-Z]+\d+)"[^>]*(?:/>|>.*?</c>)', re.DOTALL
        )
        original_cells = {
            match.group(1): match.group(0)
            for match in cell_pattern.finditer(original_sheet)
        }
        result_cells = {
            match.group(1): match.group(0)
            for match in cell_pattern.finditer(result_sheet)
        }
        for coordinate, original_cell in original_cells.items():
            if coordinate not in {b"D2", b"E2"}:
                assert result_cells[coordinate] == original_cell
        assert result_archive.read("xl/styles.xml") == original_archive.read(
            "xl/styles.xml"
        )


def test_inventory_copy_blocks_incompatible_history_headers(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet"
    worksheet.append(HEADERS)
    worksheet.append(["0001", "상품", "1정", 10, 100, "S1", "공급사"])
    history = workbook.create_sheet("입고반영내역")
    history.append(["호환되지 않는 열"])
    workbook.save(source)
    workbook.close()

    with pytest.raises(ExcelValidationError, match="기존 헤더가 호환되지"):
        create_inventory_copy(
            source,
            destination,
            [],
            [
                HistorySheetAppend(
                    sheet_name="입고반영내역",
                    headers=("작업 ID", "상품코드"),
                    rows=(("job-1", "0001"),),
                )
            ],
        )

    assert not destination.exists()


def test_inventory_copy_rechecks_product_code_before_edit(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    save_workbook(
        source,
        rows=[["0001", "상품", "1정", 10, 100, "S1", "공급사"]],
    )

    with pytest.raises(ExcelValidationError, match="상품코드가 예상값과 다릅니다"):
        create_inventory_copy(
            source,
            destination,
            [
                InventoryCellUpdate(
                    excel_row=2,
                    expected_product_code="다른상품",
                    current_stock=15,
                )
            ],
        )

    assert not destination.exists()


def test_inventory_copy_inserts_missing_target_cell_in_column_order(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "result.xlsx"
    save_workbook(
        source,
        rows=[["0001", "상품", "1정", None, 100, "S1", "공급사"]],
    )

    create_inventory_copy(
        source,
        destination,
        [
            InventoryCellUpdate(
                excel_row=2,
                expected_product_code="0001",
                current_stock=0,
            )
        ],
    )

    with ZipFile(destination) as archive:
        sheet_xml = archive.read("xl/worksheets/sheet1.xml")
    row = re.search(rb'<row\b[^>]*\br="2"[^>]*>.*?</row>', sheet_xml).group(0)
    coordinates = re.findall(rb'<c\b[^>]*\br="([A-Z]+2)"', row)
    assert coordinates == [b"A2", b"B2", b"C2", b"D2", b"E2", b"F2", b"G2"]
