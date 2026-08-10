from datetime import date
from io import BytesIO
from pathlib import Path
from threading import Barrier, Lock, Thread

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select

from app.core.config import Settings
from app.models import (
    CompletedDocument,
    Document,
    DocumentStatus,
    Item,
    Job,
    JobStatus,
    ProductIndex,
)
from app.models.job import utc_now
from app.schemas.documents import ItemUpdate
from app.services.excel import sha256_file
from app.services import exports as export_service
from app.services import items as item_service
from tests.integration.test_documents import (
    build_test_client,
    create_job,
    image_bytes,
    upload_excel,
)
from tests.integration.test_review_workflow import add_item


def complete_document(session_factory, document_id: str, transaction_date=None) -> None:
    with session_factory() as session:
        document = session.get(Document, document_id)
        document.status = DocumentStatus.COMPLETED
        document.transaction_date = transaction_date
        document.photo_supplier = "사진 공급사"
        document.invoice_number = f"INV-{document.source_order}"
        session.commit()


def add_completed_document(
    session_factory,
    tmp_path: Path,
    job_id: str,
    source_order: int,
    transaction_date,
) -> str:
    image_path = tmp_path / f"document-{source_order}.png"
    image_path.write_bytes(b"test-image")
    with session_factory() as session:
        document = Document(
            job_id=job_id,
            source_order=source_order,
            original_image_path=str(image_path),
            original_image_name=image_path.name,
            image_sha256=f"{source_order + 1:064x}",
            status=DocumentStatus.COMPLETED,
            photo_supplier="사진 공급사",
            transaction_date=transaction_date,
            invoice_number=f"INV-{source_order}",
        )
        session.add(document)
        session.commit()
        return document.id


def test_review_summary_reports_blockers_and_manual_price_resolution(
    tmp_path: Path,
) -> None:
    client, session_factory, engine = build_test_client(tmp_path)
    with client:
        job_id = create_job(client)
        empty = client.get(f"/api/jobs/{job_id}/review-summary")
        upload_excel(client, job_id)
        upload_only = client.get(f"/api/jobs/{job_id}/review-summary")
        blocked_empty_export = client.post(
            f"/api/jobs/{job_id}/export", json={"approved_by": "승인자"}
        )

        documents = [
            add_completed_document(
                session_factory, tmp_path, job_id, index, transaction_date
            )
            for index, transaction_date in enumerate(
                (date(2026, 8, 1), date(2026, 8, 3), date(2026, 8, 3))
            )
        ]
        prices = (10_000, 11_000, 12_000)
        items = [
            add_item(client, document_id, unit_price=unit_price)
            for document_id, unit_price in zip(documents, prices)
        ]
        for item in items:
            client.patch(
                f"/api/items/{item['id']}",
                json={
                    "review_status": "approved",
                    "apply_purchase_price": True,
                },
            )

        unresolved = client.get(f"/api/jobs/{job_id}/review-summary")
        invalid = client.put(
            f"/api/jobs/{job_id}/price-resolutions/0001",
            json={"selected_item_id": "missing-item"},
        )
        stale = client.put(
            f"/api/jobs/{job_id}/price-resolutions/0001",
            json={"selected_item_id": items[0]["id"]},
        )
        resolved = client.put(
            f"/api/jobs/{job_id}/price-resolutions/0001",
            json={"selected_item_id": items[2]["id"]},
        )

    assert empty.status_code == 200
    assert "ORIGINAL_EXCEL_MISSING" in {
        blocker["code"] for blocker in empty.json()["blockers"]
    }
    assert {"NO_DOCUMENTS", "NO_REVIEW_ITEMS"} <= {
        blocker["code"] for blocker in upload_only.json()["blockers"]
    }
    assert upload_only.json()["ready_to_export"] is False
    assert blocked_empty_export.status_code == 409
    assert unresolved.status_code == 200
    product = unresolved.json()["products"][0]
    assert product["price_resolution_method"] == "unresolved"
    assert product["final_purchase_price"] is None
    assert "UNRESOLVED_PRICE" in {
        blocker["code"] for blocker in unresolved.json()["blockers"]
    }
    assert invalid.status_code == 409
    assert stale.status_code == 409
    resolved_product = resolved.json()["products"][0]
    assert resolved_product["price_resolution_method"] == "manual"
    assert resolved_product["final_purchase_price"] == 12_000
    assert resolved_product["price_candidates"][2]["selected"] is True
    assert resolved.json()["ready_to_export"] is True
    engine.dispose()


def test_review_counts_only_products_with_actual_value_changes(
    tmp_path: Path,
) -> None:
    client, session_factory, engine = build_test_client(tmp_path)
    with client:
        job_id = create_job(client)
        upload_excel(client, job_id)
        document_id = add_completed_document(
            session_factory, tmp_path, job_id, 0, date(2026, 8, 3)
        )
        item = add_item(
            client,
            document_id,
            stock_increment=0,
            unit_price=1000,
        )
        approved = client.patch(
            f"/api/items/{item['id']}",
            json={"review_status": "approved"},
        )
        summary = client.get(f"/api/jobs/{job_id}/review-summary")

    assert approved.status_code == 200
    assert summary.status_code == 200
    product = summary.json()["products"][0]
    assert product["base_stock"] == product["final_stock"] == 0
    assert product["base_purchase_price"] == product["final_purchase_price"] == 1000
    assert summary.json()["counts"]["inventory_products"] == 0
    assert summary.json()["counts"]["price_products"] == 0
    engine.dispose()


def test_review_counts_price_added_to_empty_excel_price_as_change(
    tmp_path: Path,
) -> None:
    client, session_factory, engine = build_test_client(tmp_path)
    with client:
        job_id = create_job(client)
        stream = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(
            [
                "상품코드",
                "상품명",
                "규격",
                "현재고",
                "매입단가",
                "공급사코드",
                "공급사",
            ]
        )
        worksheet.append(["0001", "상품", "1정", 0, None, "S1", "공급사"])
        workbook.save(stream)
        workbook.close()
        uploaded_excel = client.post(
            f"/api/jobs/{job_id}/excel",
            files={
                "file": (
                    "products-empty-price.xlsx",
                    stream.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        document_id = add_completed_document(
            session_factory, tmp_path, job_id, 0, date(2026, 8, 3)
        )
        item = add_item(
            client,
            document_id,
            stock_increment=0,
            unit_price=1000,
        )
        approved = client.patch(
            f"/api/items/{item['id']}",
            json={"review_status": "approved"},
        )
        summary = client.get(f"/api/jobs/{job_id}/review-summary")

    assert uploaded_excel.status_code == 200
    assert approved.status_code == 200
    assert summary.status_code == 200
    product = summary.json()["products"][0]
    assert product["base_purchase_price"] is None
    assert product["final_purchase_price"] == 1000
    assert summary.json()["counts"]["price_products"] == 1
    engine.dispose()


def test_successful_export_writes_totals_histories_and_is_idempotent(
    tmp_path: Path,
) -> None:
    client, session_factory, engine = build_test_client(tmp_path)
    with client:
        job_id = create_job(client)
        upload_excel(client, job_id)
        with session_factory() as session:
            source_path = Path(session.get(Job, job_id).original_excel_path)
        original_hash = sha256_file(source_path)
        document = client.post(
            f"/api/jobs/{job_id}/documents",
            files=[("files", ("invoice.png", image_bytes(), "image/png"))],
        ).json()[0]
        complete_document(
            session_factory, document["id"], date(2026, 8, 3)
        )

        first = add_item(client, document["id"], stock_increment=2, unit_price=1200)
        second = add_item(client, document["id"], stock_increment=3, unit_price=1000)
        unchecked = add_item(
            client, document["id"], stock_increment=4, unit_price=1000
        )
        excluded = add_item(client, document["id"], matched=False)
        for item in (first, second):
            client.patch(
                f"/api/items/{item['id']}", json={"review_status": "approved"}
            )
        client.patch(
            f"/api/items/{unchecked['id']}",
            json={"review_status": "approved", "apply_inventory": False},
        )
        client.patch(
            f"/api/items/{excluded['id']}",
            json={"review_status": "excluded", "exclusion_reason": "상품 없음"},
        )

        summary = client.get(f"/api/jobs/{job_id}/review-summary")
        exported = client.post(
            f"/api/jobs/{job_id}/export", json={"approved_by": "홍길동"}
        )
        downloaded = client.get(f"/api/jobs/{job_id}/result")
        repeated = client.post(
            f"/api/jobs/{job_id}/export", json={"approved_by": "다른 승인자"}
        )
        duplicate_job_id = create_job(client)
        completed_duplicate = client.post(
            f"/api/jobs/{duplicate_job_id}/documents",
            files=[("files", ("renamed.png", image_bytes(), "image/png"))],
        ).json()[0]

    assert summary.status_code == 200
    assert summary.json()["ready_to_export"] is True, summary.json()["blockers"]
    product = summary.json()["products"][0]
    assert product["base_stock"] == 0
    assert product["stock_increment"] == 5
    assert product["final_stock"] == 5
    assert product["final_purchase_price"] == 1200
    assert set(product["item_ids"]) == {first["id"], second["id"], unchecked["id"]}
    assert exported.status_code == 200
    assert exported.json()["status"] == "completed"
    assert exported.json()["approved_by"] == "홍길동"
    assert downloaded.status_code == 200
    assert downloaded.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert repeated.status_code == 200
    assert repeated.json()["result_path"] == exported.json()["result_path"]
    assert completed_duplicate["status"] == "failed"
    assert completed_duplicate["duplicate_status"] == "confirmed"

    result_path = Path(exported.json()["result_path"])
    workbook = load_workbook(result_path, data_only=False)
    assert workbook["Sheet"]["D2"].value == 5
    assert workbook["Sheet"]["E2"].value == 1200
    assert workbook["입고반영내역"].max_row == 4
    assert workbook["검수제외내역"].max_row == 2
    history = workbook["입고반영내역"]
    headers = {
        history.cell(row=1, column=column).value: column
        for column in range(1, history.max_column + 1)
    }
    unchecked_row = next(
        row
        for row in range(2, history.max_row + 1)
        if history.cell(row=row, column=headers["품목 ID"]).value
        == unchecked["id"]
    )
    assert history.cell(
        row=unchecked_row, column=headers["재고 반영"]
    ).value is False
    assert history.cell(
        row=unchecked_row, column=headers["매입단가 반영"]
    ).value is False
    assert history.cell(
        row=unchecked_row, column=headers["변경 후 재고"]
    ).value == 5
    assert history.cell(
        row=unchecked_row, column=headers["원본 Excel SHA-256"]
    ).value == original_hash
    workbook.close()
    assert sha256_file(source_path) == original_hash
    with session_factory() as session:
        assert session.scalar(
            select(func.count(CompletedDocument.id)).where(
                CompletedDocument.job_id == job_id
            )
        ) == 1
    engine.dispose()


def test_export_appends_user_registered_product_without_changing_source(
    tmp_path: Path,
) -> None:
    client, session_factory, engine = build_test_client(tmp_path)
    with client:
        job_id = create_job(client)
        upload_excel(client, job_id)
        with session_factory() as session:
            source_path = Path(session.get(Job, job_id).original_excel_path)
        original_hash = sha256_file(source_path)
        document_id = add_completed_document(
            session_factory, tmp_path, job_id, 0, date(2026, 8, 9)
        )
        unmatched = add_item(
            client,
            document_id,
            matched=False,
            stock_increment=2,
            unit_price=1500,
        )
        registered = client.post(
            f"/api/items/{unmatched['id']}/register-product",
            json={
                "product_code": "NEW-001",
                "product_name": "신규 입고 상품",
                "specification": "30정",
                "current_stock": 4,
                "purchase_price": 1500,
                "supplier_code": "SUP-N",
                "supplier": "신규 공급사",
            },
        )
        summary = client.get(f"/api/jobs/{job_id}/review-summary")
        exported = client.post(
            f"/api/jobs/{job_id}/export", json={"approved_by": "홍길동"}
        )

    assert registered.status_code == 200
    assert summary.json()["ready_to_export"] is True, summary.json()["blockers"]
    assert exported.status_code == 200
    result_path = Path(exported.json()["result_path"])
    result = load_workbook(result_path, data_only=False)
    assert [result["Sheet"].cell(row=3, column=column).value for column in range(1, 8)] == [
        "NEW-001",
        "신규 입고 상품",
        "30정",
        6,
        1500,
        "SUP-N",
        "신규 공급사",
    ]
    result.close()
    source = load_workbook(source_path, data_only=False)
    assert source["Sheet"].max_row == 2
    source.close()
    assert sha256_file(source_path) == original_hash
    engine.dispose()


def test_export_verification_failure_rolls_back_and_removes_result(
    tmp_path: Path, monkeypatch
) -> None:
    client, session_factory, engine = build_test_client(tmp_path)
    with client:
        job_id = create_job(client)
        upload_excel(client, job_id)
        document = client.post(
            f"/api/jobs/{job_id}/documents",
            files=[("files", ("invoice.png", image_bytes(), "image/png"))],
        ).json()[0]
        complete_document(session_factory, document["id"], date(2026, 8, 3))
        item = add_item(client, document["id"], stock_increment=2, unit_price=100)
        client.patch(
            f"/api/items/{item['id']}", json={"review_status": "approved"}
        )

        def fail_verification(source_path, destination_path, updates, history_sheets):
            destination_path.parent.mkdir(parents=True, exist_ok=True)
            destination_path.write_bytes(b"partial-result")
            raise RuntimeError("출력 검증 실패")

        monkeypatch.setattr(
            "app.services.exports.create_inventory_copy", fail_verification
        )
        response = client.post(
            f"/api/jobs/{job_id}/export", json={"approved_by": "홍길동"}
        )

    assert response.status_code == 500
    with session_factory() as session:
        job = session.get(Job, job_id)
        assert job.status == JobStatus.REVIEWING
        assert job.result_path is None
        assert job.approved_by is None
        assert job.completed_at is None
        assert session.scalar(
            select(func.count(CompletedDocument.id)).where(
                CompletedDocument.job_id == job_id
            )
        ) == 0
    assert not list((tmp_path / "data" / "exports").glob("*.xlsx"))
    engine.dispose()


def test_result_statuses_and_exporting_lock_are_explicit(tmp_path: Path) -> None:
    client, session_factory, engine = build_test_client(tmp_path)
    with client:
        job_id = create_job(client)
        not_ready = client.get(f"/api/jobs/{job_id}/result")
        blank_approver = client.post(
            f"/api/jobs/{job_id}/export", json={"approved_by": "   "}
        )
        missing_excel = client.post(
            f"/api/jobs/{job_id}/export", json={"approved_by": "홍길동"}
        )
        after_missing_excel = client.get(f"/api/jobs/{job_id}")
        with session_factory() as session:
            session.get(Job, job_id).status = JobStatus.EXPORTING
            session.commit()
        exporting = client.post(
            f"/api/jobs/{job_id}/export", json={"approved_by": "홍길동"}
        )
        mutation = client.patch(
            "/api/jobs/{}/items/bulk".format(job_id),
            json={
                "target_review_status": "pending",
                "apply_inventory": False,
            },
        )

    assert not_ready.status_code == 409
    assert blank_approver.status_code == 422
    assert missing_excel.status_code == 409
    assert after_missing_excel.json()["status"] == "draft"
    assert exporting.status_code == 409
    assert mutation.status_code == 409
    engine.dispose()


def test_original_sha_and_canonical_product_row_are_revalidated(
    tmp_path: Path,
) -> None:
    client, session_factory, engine = build_test_client(tmp_path)
    with client:
        job_id = create_job(client)
        upload_excel(client, job_id)
        document_id = add_completed_document(
            session_factory, tmp_path, job_id, 0, date(2026, 8, 3)
        )
        item = add_item(client, document_id, unit_price=1000)
        client.patch(
            f"/api/items/{item['id']}", json={"review_status": "approved"}
        )
        with session_factory() as session:
            product = session.scalar(
                select(ProductIndex).where(ProductIndex.job_id == job_id)
            )
            product.excel_row = 3
            session.commit()
        invalid_row = client.get(f"/api/jobs/{job_id}/review-summary")

        with session_factory() as session:
            product = session.scalar(
                select(ProductIndex).where(ProductIndex.job_id == job_id)
            )
            product.excel_row = 2
            session.get(Job, job_id).original_excel_sha256 = "b" * 64
            session.commit()
        invalid_sha = client.get(f"/api/jobs/{job_id}/review-summary")
        blocked_export = client.post(
            f"/api/jobs/{job_id}/export", json={"approved_by": "홍길동"}
        )

    assert "INVALID_PRODUCT_BASIS" in {
        row["code"] for row in invalid_row.json()["blockers"]
    }
    assert "ORIGINAL_EXCEL_CHANGED" in {
        row["code"] for row in invalid_sha.json()["blockers"]
    }
    assert blocked_export.status_code == 409
    engine.dispose()


def test_final_database_commit_failure_removes_file_and_completion_rows(
    tmp_path: Path, monkeypatch
) -> None:
    client, session_factory, engine = build_test_client(tmp_path)
    with client:
        job_id = create_job(client)
        upload_excel(client, job_id)
        document_id = add_completed_document(
            session_factory, tmp_path, job_id, 0, date(2026, 8, 3)
        )
        item = add_item(client, document_id, unit_price=1000)
        client.patch(
            f"/api/items/{item['id']}", json={"review_status": "approved"}
        )

        def fail_commit(_db):
            raise RuntimeError("final commit failed")

        monkeypatch.setattr(
            "app.services.exports._commit_export_success", fail_commit
        )
        response = client.post(
            f"/api/jobs/{job_id}/export", json={"approved_by": "홍길동"}
        )

    assert response.status_code == 500
    with session_factory() as session:
        job = session.get(Job, job_id)
        assert job.status == JobStatus.REVIEWING
        assert job.result_path is None
        assert job.completed_at is None
        assert session.scalar(
            select(func.count(CompletedDocument.id)).where(
                CompletedDocument.job_id == job_id
            )
        ) == 0
    assert not list((tmp_path / "data" / "exports").glob("*.xlsx"))
    engine.dispose()


def test_startup_recovers_exporting_job_without_touching_completed_result(
    tmp_path: Path,
) -> None:
    client, session_factory, engine = build_test_client(tmp_path)
    exports_dir = tmp_path / "data" / "exports"
    exports_dir.mkdir(parents=True, exist_ok=True)
    orphan = exports_dir / "상품리스트_입고반영_20260808_010101.xlsx"
    completed_result = exports_dir / "상품리스트_입고반영_20260808_020202.xlsx"
    orphan.write_bytes(b"orphan")
    completed_result.write_bytes(b"completed")
    with session_factory() as session:
        interrupted = Job(
            status=JobStatus.EXPORTING,
            original_excel_path="products.xlsx",
            result_path=str(orphan),
            export_attempt_id="interrupted-attempt",
            approved_by="미확정",
            completed_at=utc_now(),
        )
        completed = Job(
            status=JobStatus.COMPLETED,
            result_path=str(completed_result),
            approved_by="확정",
            completed_at=utc_now(),
        )
        session.add_all([interrupted, completed])
        session.commit()
        interrupted_id = interrupted.id
        completed_id = completed.id

    with client:
        recovered = client.get(f"/api/jobs/{interrupted_id}")

    assert recovered.json()["status"] == "reviewing"
    assert recovered.json()["result_path"] is None
    assert recovered.json()["approved_by"] is None
    assert recovered.json()["completed_at"] is None
    assert not orphan.exists()
    assert completed_result.exists()
    with session_factory() as session:
        assert session.get(Job, interrupted_id).export_attempt_id is None
        assert session.get(Job, completed_id).status == JobStatus.COMPLETED
    engine.dispose()


def test_concurrent_exports_have_one_owner_and_one_completed_result(
    tmp_path: Path, monkeypatch
) -> None:
    client, session_factory, engine = build_test_client(tmp_path)
    with client:
        job_id = create_job(client)
        upload_excel(client, job_id)
        document_id = add_completed_document(
            session_factory, tmp_path, job_id, 0, date(2026, 8, 3)
        )
        item = add_item(client, document_id, unit_price=1000)
        client.patch(
            f"/api/items/{item['id']}", json={"review_status": "approved"}
        )

        original_create = export_service.create_inventory_copy
        call_lock = Lock()
        create_calls = 0

        def tracked_create(*args, **kwargs):
            nonlocal create_calls
            with call_lock:
                create_calls += 1
            return original_create(*args, **kwargs)

        monkeypatch.setattr(
            export_service, "create_inventory_copy", tracked_create
        )
        start = Barrier(2)
        outcomes: list[str] = []
        outcome_lock = Lock()
        settings = Settings(
            data_dir=tmp_path / "data",
            database_url=f"sqlite:///{tmp_path / 'documents.db'}",
            openai_api_key="test-key",
        )

        def run_export() -> None:
            with session_factory() as session:
                stale_job = session.get(Job, job_id)
                start.wait(timeout=5)
                try:
                    result = export_service.export_job(
                        session, stale_job, "동시 승인자", settings
                    )
                    outcome = result.status.value
                except export_service.ExportOperationError:
                    outcome = "conflict"
                with outcome_lock:
                    outcomes.append(outcome)

        threads = [Thread(target=run_export) for _ in range(2)]
        for thread in threads:
            thread.start()
        for thread in threads:
            thread.join(timeout=10)
        assert all(not thread.is_alive() for thread in threads)

    assert sorted(outcomes) == ["completed", "conflict"]
    assert create_calls == 1
    with session_factory() as session:
        job = session.get(Job, job_id)
        assert job.status == JobStatus.COMPLETED
        assert job.export_attempt_id is None
        assert job.result_path is not None
        assert Path(job.result_path).is_file()
        assert session.scalar(
            select(func.count(CompletedDocument.id)).where(
                CompletedDocument.job_id == job_id
            )
        ) == 1
    assert len(list((tmp_path / "data" / "exports").glob("*.xlsx"))) == 1
    engine.dispose()


def test_stale_mutation_is_rejected_after_export_owns_job(
    tmp_path: Path, monkeypatch
) -> None:
    client, session_factory, engine = build_test_client(tmp_path)
    with client:
        job_id = create_job(client)
        upload_excel(client, job_id)
        document_id = add_completed_document(
            session_factory, tmp_path, job_id, 0, date(2026, 8, 3)
        )
        item_read = add_item(client, document_id, stock_increment=2, unit_price=1000)
        client.patch(
            f"/api/items/{item_read['id']}",
            json={"review_status": "approved"},
        )

        export_entered = Barrier(2)
        release_export = Barrier(2)
        original_create = export_service.create_inventory_copy

        def paused_create(*args, **kwargs):
            export_entered.wait(timeout=5)
            release_export.wait(timeout=5)
            return original_create(*args, **kwargs)

        monkeypatch.setattr(
            export_service, "create_inventory_copy", paused_create
        )
        settings = Settings(
            data_dir=tmp_path / "data",
            database_url=f"sqlite:///{tmp_path / 'documents.db'}",
            openai_api_key="test-key",
        )
        export_outcomes = []

        def run_export() -> None:
            with session_factory() as session:
                result = export_service.export_job(
                    session,
                    session.get(Job, job_id),
                    "승인자",
                    settings,
                )
                export_outcomes.append(result)

        with session_factory() as stale_session:
            stale_item = stale_session.get(Item, item_read["id"])
            _ = stale_item.document.job.status
            thread = Thread(target=run_export)
            thread.start()
            export_entered.wait(timeout=5)
            try:
                item_service.update_item(
                    stale_session,
                    stale_item,
                    ItemUpdate(stock_increment=9),
                )
                mutation_outcome = "updated"
            except item_service.ItemOperationError:
                mutation_outcome = "conflict"
            release_export.wait(timeout=5)
            thread.join(timeout=10)
            assert not thread.is_alive()

    assert mutation_outcome == "conflict"
    assert len(export_outcomes) == 1
    result_path = Path(export_outcomes[0].result_path)
    workbook = load_workbook(result_path, data_only=False)
    assert workbook["Sheet"]["D2"].value == 2
    workbook.close()
    with session_factory() as session:
        assert session.get(Item, item_read["id"]).stock_increment == 2
        assert session.get(Job, job_id).status == JobStatus.COMPLETED
    engine.dispose()


def test_export_waits_for_inflight_mutation_and_uses_committed_value(
    tmp_path: Path, monkeypatch
) -> None:
    client, session_factory, engine = build_test_client(tmp_path)
    with client:
        job_id = create_job(client)
        upload_excel(client, job_id)
        document_id = add_completed_document(
            session_factory, tmp_path, job_id, 0, date(2026, 8, 3)
        )
        item_read = add_item(client, document_id, stock_increment=2, unit_price=1000)
        client.patch(
            f"/api/items/{item_read['id']}",
            json={"review_status": "approved"},
        )

        mutation_entered = Barrier(2)
        release_mutation = Barrier(2)
        export_attempted = Barrier(2)
        original_recalculate = item_service.recalculate_item_match
        original_acquire = export_service._acquire_export_lock

        def paused_recalculate(*args, **kwargs):
            mutation_entered.wait(timeout=5)
            release_mutation.wait(timeout=5)
            return original_recalculate(*args, **kwargs)

        def signaled_acquire(*args, **kwargs):
            export_attempted.wait(timeout=5)
            return original_acquire(*args, **kwargs)

        monkeypatch.setattr(
            item_service, "recalculate_item_match", paused_recalculate
        )
        monkeypatch.setattr(
            export_service, "_acquire_export_lock", signaled_acquire
        )
        settings = Settings(
            data_dir=tmp_path / "data",
            database_url=f"sqlite:///{tmp_path / 'documents.db'}",
            openai_api_key="test-key",
        )
        mutation_outcomes = []
        export_outcomes = []

        def run_mutation() -> None:
            with session_factory() as session:
                result = item_service.update_item(
                    session,
                    session.get(Item, item_read["id"]),
                    ItemUpdate(stock_increment=9),
                )
                mutation_outcomes.append(result)

        def run_export() -> None:
            with session_factory() as session:
                result = export_service.export_job(
                    session,
                    session.get(Job, job_id),
                    "승인자",
                    settings,
                )
                export_outcomes.append(result)

        mutation_thread = Thread(target=run_mutation)
        mutation_thread.start()
        mutation_entered.wait(timeout=5)
        export_thread = Thread(target=run_export)
        export_thread.start()
        export_attempted.wait(timeout=5)
        release_mutation.wait(timeout=5)
        mutation_thread.join(timeout=10)
        export_thread.join(timeout=10)
        assert not mutation_thread.is_alive()
        assert not export_thread.is_alive()

    assert len(mutation_outcomes) == 1
    assert len(export_outcomes) == 1
    result_path = Path(export_outcomes[0].result_path)
    workbook = load_workbook(result_path, data_only=False)
    assert workbook["Sheet"]["D2"].value == 9
    workbook.close()
    with session_factory() as session:
        assert session.get(Item, item_read["id"]).stock_increment == 9
        assert session.get(Job, job_id).status == JobStatus.COMPLETED
    engine.dispose()


def test_parallel_different_jobs_use_distinct_result_paths(
    tmp_path: Path, monkeypatch
) -> None:
    client, session_factory, engine = build_test_client(tmp_path)
    with client:
        jobs: list[tuple[str, int]] = []
        for source_order, increment in enumerate((2, 7)):
            job_id = create_job(client)
            upload_excel(client, job_id)
            document_id = add_completed_document(
                session_factory,
                tmp_path,
                job_id,
                source_order,
                date(2026, 8, 3),
            )
            item = add_item(
                client,
                document_id,
                stock_increment=increment,
                unit_price=1000,
            )
            client.patch(
                f"/api/items/{item['id']}",
                json={"review_status": "approved"},
            )
            jobs.append((job_id, increment))

        create_barrier = Barrier(2)
        original_create = export_service.create_inventory_copy

        def synchronized_create(*args, **kwargs):
            create_barrier.wait(timeout=5)
            return original_create(*args, **kwargs)

        monkeypatch.setattr(
            export_service, "create_inventory_copy", synchronized_create
        )
        settings = Settings(
            data_dir=tmp_path / "data",
            database_url=f"sqlite:///{tmp_path / 'documents.db'}",
            openai_api_key="test-key",
        )
        outcomes = {}
        outcomes_lock = Lock()

        def run_export(job_id: str) -> None:
            with session_factory() as session:
                result = export_service.export_job(
                    session,
                    session.get(Job, job_id),
                    "승인자",
                    settings,
                )
                with outcomes_lock:
                    outcomes[job_id] = result

        threads = [Thread(target=run_export, args=(job_id,)) for job_id, _ in jobs]
        for thread in threads:
            thread.start()
        for thread in threads:
            thread.join(timeout=10)
        assert all(not thread.is_alive() for thread in threads)

    assert set(outcomes) == {job_id for job_id, _ in jobs}
    paths = {Path(result.result_path) for result in outcomes.values()}
    assert len(paths) == 2
    assert all(path.is_file() for path in paths)
    for job_id, increment in jobs:
        workbook = load_workbook(outcomes[job_id].result_path, data_only=False)
        assert workbook["Sheet"]["D2"].value == increment
        workbook.close()
    assert len(list((tmp_path / "data" / "exports").glob("*.xlsx"))) == 2
    engine.dispose()


def test_parallel_duplicate_documents_finalize_only_one_job(
    tmp_path: Path, monkeypatch
) -> None:
    client, session_factory, engine = build_test_client(tmp_path)
    with client:
        job_ids = []
        for _ in range(2):
            job_id = create_job(client)
            upload_excel(client, job_id)
            document_id = add_completed_document(
                session_factory, tmp_path, job_id, 0, date(2026, 8, 3)
            )
            item = add_item(
                client,
                document_id,
                stock_increment=2,
                unit_price=1000,
            )
            client.patch(
                f"/api/items/{item['id']}",
                json={"review_status": "approved"},
            )
            job_ids.append(job_id)

        create_barrier = Barrier(2)
        original_create = export_service.create_inventory_copy
        create_calls = 0
        create_lock = Lock()

        def synchronized_create(*args, **kwargs):
            nonlocal create_calls
            with create_lock:
                create_calls += 1
            create_barrier.wait(timeout=5)
            return original_create(*args, **kwargs)

        monkeypatch.setattr(
            export_service, "create_inventory_copy", synchronized_create
        )
        settings = Settings(
            data_dir=tmp_path / "data",
            database_url=f"sqlite:///{tmp_path / 'documents.db'}",
            openai_api_key="test-key",
        )
        outcomes = {}
        outcomes_lock = Lock()

        def run_export(job_id: str) -> None:
            with session_factory() as session:
                try:
                    result = export_service.export_job(
                        session,
                        session.get(Job, job_id),
                        "승인자",
                        settings,
                    )
                    outcome = ("completed", result.result_path)
                except export_service.ExportOperationError:
                    outcome = ("conflict", None)
                with outcomes_lock:
                    outcomes[job_id] = outcome

        threads = [Thread(target=run_export, args=(job_id,)) for job_id in job_ids]
        for thread in threads:
            thread.start()
        for thread in threads:
            thread.join(timeout=10)
        assert all(not thread.is_alive() for thread in threads)

    assert create_calls == 2
    assert sorted(outcome[0] for outcome in outcomes.values()) == [
        "completed",
        "conflict",
    ]
    completed_path = Path(
        next(path for status, path in outcomes.values() if status == "completed")
    )
    assert completed_path.is_file()
    with session_factory() as session:
        statuses = [session.get(Job, job_id).status for job_id in job_ids]
        assert sorted(status.value for status in statuses) == [
            "completed",
            "reviewing",
        ]
        assert session.scalar(select(func.count(CompletedDocument.id))) == 1
    assert list((tmp_path / "data" / "exports").glob("*.xlsx")) == [
        completed_path
    ]
    engine.dispose()


def test_excel_replacement_is_blocked_and_preserves_approved_review_data(
    tmp_path: Path,
) -> None:
    client, session_factory, engine = build_test_client(tmp_path)
    with client:
        job_id = create_job(client)
        upload_excel(client, job_id)
        document_id = add_completed_document(
            session_factory, tmp_path, job_id, 0, date(2026, 8, 3)
        )
        item = add_item(client, document_id, stock_increment=2, unit_price=1000)
        client.patch(
            f"/api/items/{item['id']}",
            json={"review_status": "approved"},
        )

        stream = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(
            [
                "상품코드",
                "상품명",
                "규격",
                "현재고",
                "매입단가",
                "공급사코드",
                "공급사",
            ]
        )
        worksheet.append(
            ["0001", "변경 상품", "2정", 0, 2000, "S2", "변경 공급사"]
        )
        workbook.save(stream)
        workbook.close()
        replaced = client.post(
            f"/api/jobs/{job_id}/excel",
            files={
                "file": (
                    "products-updated.xlsx",
                    stream.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        persisted_items = client.get(f"/api/jobs/{job_id}/items")
        summary = client.get(f"/api/jobs/{job_id}/review-summary")
        exported = client.post(
            f"/api/jobs/{job_id}/export", json={"approved_by": "승인자"}
        )

    assert replaced.status_code == 409
    assert "추출 또는 검수 데이터" in replaced.json()["detail"]
    persisted_item = persisted_items.json()[0]
    assert persisted_item["matched_product_name"] == "상품"
    assert persisted_item["matched_supplier"] == "공급사"
    assert persisted_item["base_purchase_price"] == 1000
    assert persisted_item["review_status"] == "approved"
    assert summary.status_code == 200
    assert summary.json()["ready_to_export"] is True, summary.json()["blockers"]
    assert summary.json()["products"][0]["base_stock"] == 0
    assert summary.json()["products"][0]["final_stock"] == 2
    assert summary.json()["products"][0]["base_purchase_price"] == 1000
    assert summary.json()["products"][0]["final_purchase_price"] == 1000
    assert exported.status_code == 200
    result = load_workbook(exported.json()["result_path"], data_only=False)
    assert result["Sheet"]["D2"].value == 2
    assert result["Sheet"]["E2"].value == 1000
    assert result["검수제외내역"].max_row == 1
    result.close()
    engine.dispose()
