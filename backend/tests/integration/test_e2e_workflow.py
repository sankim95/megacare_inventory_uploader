from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.services.extraction import AIExtractionError
from tests.integration.test_documents import (
    build_test_client,
    create_job,
    image_bytes,
    parsed_invoice,
    upload_excel,
)


def test_complete_inventory_workflow_through_http_api(
    tmp_path: Path, monkeypatch
) -> None:
    client, _, engine = build_test_client(tmp_path)
    first_image = image_bytes()
    second_image = image_bytes("JPEG")
    extraction_calls = []

    def invoice(name: str, invoice_number: str, transaction_date: date):
        result = parsed_invoice(name)
        result.document.invoice_number = invoice_number
        result.document.transaction_date = transaction_date
        return result

    def first_extraction_pass(**kwargs):
        extraction_calls.append(kwargs)
        if len(extraction_calls) == 2:
            raise AIExtractionError("temporary test failure")
        return invoice("첫 번째 상품", "INV-E2E-1", date(2026, 8, 6))

    monkeypatch.setattr(
        "app.services.extraction.parse_invoice_image", first_extraction_pass
    )

    with client:
        job_id = create_job(client)
        upload_excel(client, job_id)
        uploaded = client.post(
            f"/api/jobs/{job_id}/documents",
            files=[
                ("files", ("first.png", first_image, "image/png")),
                ("files", ("second.jpg", second_image, "image/jpeg")),
            ],
        )
        extracted = client.post(f"/api/jobs/{job_id}/extract")

        monkeypatch.setattr(
            "app.services.extraction.parse_invoice_image",
            lambda **_: invoice(
                "재시도 상품", "INV-E2E-2", date(2026, 8, 7)
            ),
        )
        retried = client.post(
            f"/api/documents/{uploaded.json()[1]['id']}/extract"
        )
        matched = client.post(f"/api/jobs/{job_id}/match")
        matched_items = matched.json()

        corrected = client.patch(
            f"/api/items/{matched_items[0]['id']}",
            json={
                "stock_increment": 4,
                "unit_price": 1200,
                "amount": 2400,
                "apply_purchase_price": True,
                "notes": "E2E 수량·단가 확인",
            },
        )
        approved = client.patch(
            f"/api/items/{matched_items[0]['id']}",
            json={"review_status": "approved"},
        )
        excluded = client.patch(
            f"/api/items/{matched_items[1]['id']}",
            json={
                "review_status": "excluded",
                "exclusion_reason": "검수 결과 반영 제외",
            },
        )
        summary = client.get(f"/api/jobs/{job_id}/review-summary")
        exported = client.post(
            f"/api/jobs/{job_id}/export", json={"approved_by": "E2E 검수자"}
        )
        downloaded = client.get(f"/api/jobs/{job_id}/result")

        duplicate_job_id = create_job(client)
        duplicate_upload = client.post(
            f"/api/jobs/{duplicate_job_id}/documents",
            files=[
                (
                    "files",
                    ("renamed-after-export.png", first_image, "image/png"),
                )
            ],
        )

    assert uploaded.status_code == 200
    assert [row["status"] for row in uploaded.json()] == ["pending", "pending"]
    assert extracted.status_code == 200
    assert [row["status"] for row in extracted.json()] == [
        "completed",
        "failed",
    ]
    assert len(extraction_calls) == 2
    assert all(call["api_key"] == "test-key" for call in extraction_calls)
    assert all(call["model"] == "gpt-test" for call in extraction_calls)
    assert retried.status_code == 200
    assert retried.json()["status"] == "completed"
    assert matched.status_code == 200
    assert [row["matched_product_code"] for row in matched_items] == [
        "0001",
        "0001",
    ]
    assert corrected.status_code == 200
    assert corrected.json()["stock_increment"] == 4
    assert corrected.json()["unit_price"] == 1200
    assert corrected.json()["apply_purchase_price"] is True
    assert approved.status_code == 200
    assert approved.json()["review_status"] == "approved"
    assert excluded.status_code == 200
    assert excluded.json()["review_status"] == "excluded"

    assert summary.status_code == 200
    assert summary.json()["ready_to_export"] is True, summary.json()["blockers"]
    assert summary.json()["counts"] == {
        "approved_items": 1,
        "excluded_items": 1,
        "pending_items": 0,
        "inventory_products": 1,
        "price_products": 1,
    }
    product = summary.json()["products"][0]
    assert product["base_stock"] == 0
    assert product["stock_increment"] == 4
    assert product["final_stock"] == 4
    assert product["base_purchase_price"] == 1000
    assert product["final_purchase_price"] == 1200

    assert exported.status_code == 200
    assert exported.json()["status"] == "completed"
    assert exported.json()["approved_by"] == "E2E 검수자"
    assert downloaded.status_code == 200
    result = load_workbook(BytesIO(downloaded.content), data_only=False)
    assert result["Sheet"]["D2"].value == 4
    assert result["Sheet"]["E2"].value == 1200
    assert result["입고반영내역"].max_row == 2
    assert result["검수제외내역"].max_row == 2
    result.close()

    assert duplicate_upload.status_code == 200
    duplicate = duplicate_upload.json()[0]
    assert duplicate["status"] == "failed"
    assert duplicate["duplicate_status"] == "confirmed"
    assert "동일 이미지" in duplicate["processing_error"]
    engine.dispose()
